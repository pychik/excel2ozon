import logging as logger
from openpyxl import load_workbook
from sys import exit as s_exit
from time import sleep
from config import settings


logger.basicConfig(level=logger.INFO, format="%(asctime)s %(levelname)s %(message)s")


class PriceReader:
    def __init__(self, filename: str = settings.PRICE_TABLE, start_rows: int = settings.START_ROW,
                 article_col: str = settings.ARTICLE_COLUMN, prices_delta_col: str = settings.PRICE_COLUMN,
                 delivery_col: str = settings.DELIVERY_COLUMN):

        self.ac = article_col
        self.pc = prices_delta_col
        self.dc = delivery_col
        self.start_rows = start_rows
        self.filename = filename

    def get_prices_dict(self) -> dict:
        try:
            dataframe = load_workbook(filename=self.filename).active
        except:
            logger.warning(msg=f"В папке должен находиться файл с настройками для цен {settings.TABLE_NAME}."
                               f"Отключаюсь")
            sleep(4)
            s_exit()
        # Define variable to read sheet
        # dataframe1 = dataframe.active

        # Iterate the loop to read the cell values
        # print(dataframe["B1"].value, dataframe["D1"].value)
        prices_dict = {}
        for row in range(self.start_rows, dataframe.max_row+1):
            prices_dict[dataframe[f"{self.ac}{row}"].value] = (round(float(dataframe[f"{self.pc}{row}"].value), 4), round(float(dataframe[f"{self.dc}{row}"].value), 4))
            # for col in f"{self.ac}{self.pc}":
            #     print(dataframe[f"{col}{row}"].value)
        # print(prices_dict['442511'])

        return prices_dict

    @staticmethod
    def price_process(price: float, prices_tuple: tuple[float, int]) -> None|str:
        if prices_tuple and len(prices_tuple) == 2:
            return str(round(price*((100+prices_tuple[0])/100) + prices_tuple[1]))
        else:
            return None


if __name__ == '__main__':
    pr = PriceReader()
    prices_dict = pr.get_prices_dict()
    print(PriceReader.price_process(price=135.54646, prices_tuple=prices_dict.get('НС-1155801')))