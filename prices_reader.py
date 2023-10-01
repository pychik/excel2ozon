from openpyxl import load_workbook
from config import settings
# Define variable to load the dataframe



class PriceReader:
    def __init__(self, filename: str = settings.Prices.TABLEFILE, start_rows: int = settings.Prices.START_ROW,
                 article_col: str = settings.Prices.AC, prices_delta_col: str = settings.Prices.PC):

        self.ac = article_col
        self.pc = prices_delta_col
        self.start_rows = start_rows
        self.filename = filename

    def get_prices_dict(self) -> dict:
        dataframe = load_workbook(filename=self.filename).active

        # Define variable to read sheet
        # dataframe1 = dataframe.active

        # Iterate the loop to read the cell values
        # print(dataframe["B1"].value, dataframe["D1"].value)
        prices_dict = {}
        for row in range(self.start_rows, dataframe.max_row+1):
            prices_dict[dataframe[f"{self.ac}{row}"].value] = round(float(dataframe[f"{self.pc}{row}"].value), 4 )
            # for col in f"{self.ac}{self.pc}":
            #     print(dataframe[f"{col}{row}"].value)
        # print(prices_dict['442511'])

        return prices_dict

    @staticmethod
    def price_process(price: float, price_delta: float) -> str:
        if price == 0:
            return str(settings.Prices.DELIVERY)
        else:
            return str(int(price*((100+price_delta)/100) + settings.Prices.DELIVERY))


if __name__ == '__main__':
    pr = PriceReader()
    prices_dict = pr.get_prices_dict()
    print(PriceReader.price_process(price=100, price_delta=prices_dict.get('452001')))
